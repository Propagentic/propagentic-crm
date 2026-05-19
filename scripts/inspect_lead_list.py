"""Quick read of the two lead-list XLSX files: list sheets, columns, row count, and a few sample rows.

Run from project root:
    .venv/bin/python3 scripts/inspect_lead_list.py
"""
from pathlib import Path
import openpyxl

FILES = [
    Path("/Users/jpmacair/Desktop/Propagentic Lead List 1.xlsx"),
    Path("/Users/jpmacair/Desktop/Propagentic_Lead_List_Nashville (1).xlsx"),
]

for path in FILES:
    print("=" * 80)
    print(f"FILE: {path.name}")
    print("=" * 80)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ERROR opening: {e}")
        continue

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name!r}")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  (empty)")
            continue
        header = rows[0]
        body = rows[1:]
        print(f"  rows: {len(body)} (excluding header)")
        print(f"  cols: {len(header)}")
        print(f"  columns:")
        for i, h in enumerate(header):
            print(f"    [{i:2d}] {h!r}")
        print(f"  first 3 rows:")
        for r in body[:3]:
            for i, (h, v) in enumerate(zip(header, r)):
                print(f"    {h}: {v!r}")
            print("    ---")
    print()
