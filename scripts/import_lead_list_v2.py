"""v2.0 lead-list importer.

Reads an XLSX lead list, normalizes fields, geocodes property addresses via
Google Maps, and upserts each row into the Firestore `leads` collection.

Doc IDs are a compound natural key: `{state_lower}__{zip5}__{addr_hash12}`,
so re-importing the same XLSX is idempotent — and re-importing an updated
version of the same lead refreshes source fields while preserving CRM state
(status, notes, last_touched, touched_by) that the team has added.

Usage:
    .venv/bin/python3 scripts/import_lead_list_v2.py \
        --file "/Users/jpmacair/Desktop/Propagentic Lead List 1.xlsx" \
        [--limit N]          # debug: only process first N rows
        [--no-geocode]       # skip geocoding (lat/lon will be missing)
        [--dry-run]          # parse + geocode but skip Firestore writes
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
import requests
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

ROOT = Path(__file__).resolve().parent.parent
KEY_PATH = ROOT / "secrets" / "firebase-key.json"
MAPS_KEY_PATH = ROOT / "secrets" / "google-maps-key.txt"
GEOCODE_CACHE_PATH = ROOT / "data" / "geocode_cache.json"

GEOCODING_URL = "https://maps.googleapis.com/maps/api/geocode/json"
BATCH_SIZE = 400  # Firestore batched-write limit is 500; leave headroom.

# --- Source column map (XLSX header → schema field) -----------------------
COLMAP = {
    "Parcel Id":          "external_parcel_id",
    "FullName":           "owner_name_raw",
    "FirstName":          "first_name",
    "LastName":           "last_name",
    "O2FIRST":            "co_owner_first",
    "O2LAST":             "co_owner_last",
    "Property ADDRESS":   "property_address",
    "Property ADDRESS2":  "property_address2",
    "Property CITY":      "property_city",
    "Property State":     "property_state",
    "Property Zip":       "property_zip",
    "Property COUNTY":    "property_county",
    "MAILING Address":    "mailing_address",
    "MAILING ADDRESS2":   "mailing_address2",
    "MAILING CITY":       "mailing_city",
    "MAILING STATE":      "mailing_state",
    "MAILING Zip":        "mailing_zip",
    "Land sqft":          "land_sqft",
    "OWNERTYPE":          "owner_type",
    "Total Living area":  "living_area",
    "Legal Description":  "legal_description",
    "Absntee Flag":       "absentee",
    "Vacant":             "vacant",
    "Propety Data Type":  "property_type",
    "YearBuilt":          "year_built",
    "Seller Propensity":  "seller_propensity",
    "Solar Intenders":    "solar_intender",
    "Windows_Intenders":  "windows_intender",
    "Roofing_Intenders":  "roofing_intender",
    "Phone":              "phone",
    "PhoneType":          "phone_type",
    "DNC":                "dnc",
    "Email":              "email",
}


# --- Helpers --------------------------------------------------------------
def clean_str(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None


def upper_county(v):
    """Normalize county names: strip, uppercase, replace dashes with spaces.
    Examples: 'baltimore-county' -> 'BALTIMORE', 'baltimore' -> 'BALTIMORE'."""
    if v is None: return None
    s = str(v).strip().upper()
    s = s.replace("-COUNTY", "").replace("-CITY", " CITY")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def upper_state(v):
    if v is None: return None
    s = str(v).strip().upper()
    return s[:2] if s else None


def clean_zip(v):
    if v is None: return None
    digits = re.sub(r"[^0-9]", "", str(v))
    return digits[:5] if digits else None


def yn_to_bool(v):
    if v is None: return None
    s = str(v).strip().upper()
    if s in ("Y", "YES", "TRUE", "T", "1"): return True
    if s in ("N", "NO", "FALSE", "F", "0"): return False
    return None


def safe_int(v):
    if v is None: return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def clean_phone(v):
    if v is None: return None
    digits = re.sub(r"[^0-9]", "", str(v))
    return digits if len(digits) >= 10 else None


def normalize_addr_for_id(addr, city, state, zip5):
    """Stable normalization just for doc-ID hashing. Decoupled from display."""
    parts = []
    if addr: parts.append(str(addr).upper())
    if city: parts.append(str(city).upper())
    if state: parts.append(str(state).upper())
    if zip5: parts.append(str(zip5))
    s = " ".join(parts)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_doc_id(state, zip5, addr_norm):
    h = hashlib.sha256(addr_norm.encode()).hexdigest()[:12]
    state_part = (state or "xx").lower()
    zip_part = zip5 or "00000"
    return f"{state_part}__{zip_part}__{h}"


def name_titlecase(s):
    if not s: return None
    return " ".join(p.capitalize() for p in str(s).strip().split())


# --- Geocoder -------------------------------------------------------------
class Geocoder:
    def __init__(self, api_key, cache_path):
        self.api_key = api_key
        self.cache_path = cache_path
        self.cache = {}
        if cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text())
            except Exception:
                self.cache = {}
        self.hits = 0
        self.misses = 0
        self.errors = 0

    def geocode(self, addr, city, state, zip5):
        if not addr or not state:
            return None
        # Build a clean address string and use it as cache key.
        bits = [addr, city, state, zip5]
        full = ", ".join([b for b in bits if b])
        cache_key = full.upper()

        if cache_key in self.cache:
            self.hits += 1
            return self.cache[cache_key]

        self.misses += 1
        try:
            resp = requests.get(GEOCODING_URL,
                                params={"address": full, "key": self.api_key},
                                timeout=15)
            body = resp.json()
        except Exception as e:
            self.errors += 1
            return None

        status = body.get("status")
        results = body.get("results", [])
        if status == "OK" and results:
            loc = results[0]["geometry"]["location"]
            entry = {
                "lat": loc["lat"],
                "lon": loc["lng"],
                "status": status,
                "formatted": results[0].get("formatted_address"),
                "location_type": results[0]["geometry"].get("location_type"),
            }
        else:
            entry = {"lat": None, "lon": None, "status": status}

        self.cache[cache_key] = entry
        return entry

    def save_cache(self):
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.cache_path.write_text(json.dumps(self.cache, indent=2))


# --- Row → document -------------------------------------------------------
def row_to_lead(row_dict, geocode_entry, source_file, row_idx, importer_email):
    state = upper_state(row_dict.get("property_state"))
    zip5 = clean_zip(row_dict.get("property_zip"))
    county = upper_county(row_dict.get("property_county"))
    city = clean_str(row_dict.get("property_city"))
    addr = clean_str(row_dict.get("property_address"))

    addr_norm = normalize_addr_for_id(addr, city, state, zip5)
    if not addr_norm:
        return None, None  # skip rows without a property address

    doc_id = build_doc_id(state, zip5, addr_norm)

    lat = lon = geocode_status = None
    geocode_formatted = None
    if geocode_entry:
        lat = geocode_entry.get("lat")
        lon = geocode_entry.get("lon")
        geocode_status = geocode_entry.get("status")
        geocode_formatted = geocode_entry.get("formatted")

    first = name_titlecase(row_dict.get("first_name"))
    last = name_titlecase(row_dict.get("last_name"))
    full = name_titlecase(row_dict.get("owner_name_raw")) or (
        f"{first} {last}" if first and last else first or last
    )

    doc = {
        "external_parcel_id": clean_str(row_dict.get("external_parcel_id")),
        "owner_name": full,
        "owner_name_raw": clean_str(row_dict.get("owner_name_raw")),
        "first_name": first,
        "last_name": last,
        "co_owner_first": name_titlecase(row_dict.get("co_owner_first")),
        "co_owner_last": name_titlecase(row_dict.get("co_owner_last")),

        "property_address": addr,
        "property_address2": clean_str(row_dict.get("property_address2")),
        "property_city": city,
        "property_state": state,
        "property_zip": zip5,
        "property_county": county,
        "property_lat": lat,
        "property_lon": lon,
        "geocode_status": geocode_status,
        "geocode_formatted": geocode_formatted,

        "property_type": clean_str(row_dict.get("property_type")),
        "year_built": safe_int(row_dict.get("year_built")),
        "land_sqft": safe_int(row_dict.get("land_sqft")),
        "living_area": safe_int(row_dict.get("living_area")),
        "legal_description": clean_str(row_dict.get("legal_description")),

        "mailing_address": clean_str(row_dict.get("mailing_address")),
        "mailing_address2": clean_str(row_dict.get("mailing_address2")),
        "mailing_city": clean_str(row_dict.get("mailing_city")),
        "mailing_state": upper_state(row_dict.get("mailing_state")),
        "mailing_zip": clean_zip(row_dict.get("mailing_zip")),

        "owner_type": clean_str(row_dict.get("owner_type")),
        "absentee": yn_to_bool(row_dict.get("absentee")),
        "vacant": yn_to_bool(row_dict.get("vacant")),

        "seller_propensity": clean_str(row_dict.get("seller_propensity")),
        "solar_intender": clean_str(row_dict.get("solar_intender")),
        "windows_intender": clean_str(row_dict.get("windows_intender")),
        "roofing_intender": clean_str(row_dict.get("roofing_intender")),

        "phone": clean_phone(row_dict.get("phone")),
        "phone_type": clean_str(row_dict.get("phone_type")),
        "dnc": yn_to_bool(row_dict.get("dnc")),
        "email": clean_str(row_dict.get("email")),

        "market_state": state,
        "market_county": county,
        "market": f"{(county or 'unknown').lower().replace(' ', '_')}_{(state or 'xx').lower()}",

        "source_file": source_file,
        "source_row": row_idx,
        "imported_at": firestore.SERVER_TIMESTAMP,
        "imported_by": importer_email,
    }

    # Drop None values to keep documents tidy and avoid clobbering with nulls
    # under set(merge=True). (Firestore stores 'null' values as a real null,
    # which counts as a write of None over whatever was there.)
    doc = {k: v for k, v in doc.items() if v is not None}
    return doc_id, doc


# --- Main -----------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Path to XLSX lead list")
    ap.add_argument("--sheet", default=None, help="Sheet name (default: first)")
    ap.add_argument("--limit", type=int, default=None,
                    help="Debug: only process first N rows")
    ap.add_argument("--no-geocode", action="store_true",
                    help="Skip geocoding step")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse + geocode but skip Firestore writes")
    ap.add_argument("--importer", default="jacksontracey203@propagenticai.com",
                    help="Email recorded as imported_by")
    args = ap.parse_args()

    src_path = Path(args.file).expanduser().resolve()
    if not src_path.exists():
        print(f"ERROR: file not found: {src_path}", file=sys.stderr)
        sys.exit(1)

    # --- Firestore init ---
    if not args.dry_run:
        if not KEY_PATH.exists():
            print(f"ERROR: firebase key missing: {KEY_PATH}", file=sys.stderr)
            sys.exit(1)
        cred = credentials.Certificate(str(KEY_PATH))
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        db = firestore.client()
    else:
        db = None

    # --- Geocoder init ---
    geocoder = None
    if not args.no_geocode:
        if not MAPS_KEY_PATH.exists():
            print(f"ERROR: google maps key missing: {MAPS_KEY_PATH}", file=sys.stderr)
            sys.exit(1)
        maps_key = MAPS_KEY_PATH.read_text().strip()
        geocoder = Geocoder(maps_key, GEOCODE_CACHE_PATH)
        print(f"geocoder ready; cache has {len(geocoder.cache)} entries\n")

    # --- Read XLSX ---
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("empty sheet", file=sys.stderr); sys.exit(1)
    header = rows[0]
    body = rows[1:]
    if args.limit:
        body = body[:args.limit]
    print(f"{src_path.name} :: sheet={sheet_name} :: {len(body)} rows to process\n")

    # Index headers → COLMAP target field
    col_idx = {}
    for i, h in enumerate(header):
        if h in COLMAP:
            col_idx[COLMAP[h]] = i
    missing = set(COLMAP.values()) - set(col_idx.keys())
    if missing:
        print(f"NOTE: source missing columns -> {sorted(missing)}\n")

    # --- Process rows ---
    leads_to_write = []  # list of (doc_id, dict)
    skipped_no_addr = 0
    by_market = Counter()
    started = time.time()

    for ridx, row in enumerate(body, start=2):  # row 1 is header
        row_dict = {field: row[i] for field, i in col_idx.items()}
        # Pre-extract address for geocoding lookup.
        addr = clean_str(row_dict.get("property_address"))
        city = clean_str(row_dict.get("property_city"))
        state = upper_state(row_dict.get("property_state"))
        zip5 = clean_zip(row_dict.get("property_zip"))
        if not addr:
            skipped_no_addr += 1
            continue

        geo = geocoder.geocode(addr, city, state, zip5) if geocoder else None
        doc_id, doc = row_to_lead(
            row_dict, geo, src_path.name, ridx, args.importer
        )
        if doc is None:
            skipped_no_addr += 1
            continue
        leads_to_write.append((doc_id, doc))
        by_market[(state, upper_county(row_dict.get("property_county")))] += 1

        if (len(leads_to_write) % 200) == 0:
            elapsed = time.time() - started
            print(f"  parsed {len(leads_to_write)} ({elapsed:.1f}s elapsed)")

    elapsed = time.time() - started
    print(f"\nparsing done :: {len(leads_to_write)} leads ready, "
          f"{skipped_no_addr} skipped, {elapsed:.1f}s elapsed")

    if geocoder:
        print(f"geocoder :: cache_hits={geocoder.hits} api_calls={geocoder.misses} errors={geocoder.errors}")
        geocoder.save_cache()
        print(f"cache written to {GEOCODE_CACHE_PATH}")

    print("\nmarkets:")
    for (st, co), n in sorted(by_market.items(), key=lambda x: -x[1]):
        print(f"  {st}  {co!r:>30}  {n:>5}")

    if args.dry_run:
        print("\n--dry-run: skipping Firestore writes.")
        return

    # --- Upsert to Firestore (merge=True preserves crm_* fields not in payload) ---
    print(f"\nupserting {len(leads_to_write)} docs to firestore (batched)...")
    batch = db.batch()
    n_in_batch = 0
    n_total = 0
    started = time.time()
    for doc_id, doc in leads_to_write:
        ref = db.collection("leads").document(doc_id)
        batch.set(ref, doc, merge=True)
        n_in_batch += 1
        n_total += 1
        if n_in_batch >= BATCH_SIZE:
            batch.commit()
            print(f"  ...{n_total} written")
            batch = db.batch()
            n_in_batch = 0
    if n_in_batch > 0:
        batch.commit()
    print(f"\nwrote {n_total} leads in {time.time()-started:.1f}s.")

    # --- Audit log ---
    audit = {
        "imported_at": firestore.SERVER_TIMESTAMP,
        "imported_by": args.importer,
        "source_file": src_path.name,
        "rows_in_source": len(body),
        "rows_imported": n_total,
        "rows_skipped_no_address": skipped_no_addr,
        "geocode_cache_hits": geocoder.hits if geocoder else None,
        "geocode_api_calls": geocoder.misses if geocoder else None,
        "geocode_errors": geocoder.errors if geocoder else None,
    }
    db.collection("import_log").add(audit)
    print("audit row written to import_log collection.\n")
    print("DONE.")


if __name__ == "__main__":
    main()
